import configparser
import os

import customtkinter as ctk
from openpyxl import load_workbook
from openpyxl.styles import Font


class SettingsWindow(ctk.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.config = configparser.ConfigParser()
        self.config.read('config.ini')
        self.filepath = self.config.get('path', 'path')

        self.params = {'padx': 8, 'pady': 8, 'sticky': 'nsew'}
        self.geometry("600x500")
        self.title('Ustawienia')

        self.wb_number = load_workbook(filename='rejestr.xlsx')
        self.ws_number = self.wb_number.active
        self.register_number = self.ws_number[f'B{self.ws_number.max_row}'].value

        self.name_label = ctk.CTkLabel(self, text='Nazwa firmy')
        self.name_label.grid(row=0, column=0, **self.params)

        self.name_entry = ctk.CTkEntry(self, placeholder_text='Nazwa Firmy')
        self.name_entry.grid(row=0, column=1, columnspan=3, **self.params)
        self.name_entry.insert(ctk.END, self.config.get('contractor_details', 'name'))

        self.address_label = ctk.CTkLabel(self, text='Adres firmy')
        self.address_label.grid(row=1, column=0, **self.params)

        self.zip_code_and_city_entry = ctk.CTkEntry(self, width=200, placeholder_text='Kod pocztowy i miasto')
        self.zip_code_and_city_entry.grid(row=1, column=1, **self.params)
        self.zip_code_and_city_entry.insert(ctk.END, self.config.get('contractor_details', 'zip_code_and_city'))

        self.street_and_number_entry = ctk.CTkEntry(self, width=200, placeholder_text='Ulica i numer')
        self.street_and_number_entry.grid(row=1, column=2, **self.params)
        self.street_and_number_entry.insert(ctk.END, self.config.get('contractor_details', 'street_and_number'))

        self.measuring_tools_label = ctk.CTkLabel(self, text='Narzędzia pomiarów')
        self.measuring_tools_label.grid(row=2, column=0, **self.params)

        self.measuring_tools_entry = ctk.CTkEntry(self, placeholder_text='np. MIC-3 nr 345758')
        self.measuring_tools_entry.grid(row=2, column=1, **self.params, columnspan=3)
        self.measuring_tools_entry.insert(ctk.END, self.config.get('contractor_details', 'measuring_tools'))

        self.register_number_label = ctk.CTkLabel(self, text='Aktualny numer rejestru')
        self.register_number_label.grid(row=3, column=0, **self.params)

        self.register_number_entry = ctk.CTkEntry(self)
        self.register_number_entry.grid(row=3, column=1, **self.params)
        self.register_number_entry.insert(ctk.END, ('' if self.register_number == 'Numer' else self.register_number))

        self.filepath_button = ctk.CTkButton(self, text='Ustaw ścieżkę zapisu', command=self.set_path)
        self.filepath_button.grid(row=4, column=0, **self.params)

        self.filepath_label = ctk.CTkLabel(self, text=self.filepath)
        self.filepath_label.grid(row=4, column=1, columnspan=3, **self.params)

        self.set_and_save_button = ctk.CTkButton(self, text='Zapisz dane', command=self.save_data)
        self.set_and_save_button.grid(row=5, column=2, **self.params)

    def save_data(self):
        self.config.set('contractor_details', 'name', self.name_entry.get())
        self.config.set('contractor_details', 'zip_code_and_city', self.zip_code_and_city_entry.get())
        self.config.set('contractor_details', 'street_and_number', self.street_and_number_entry.get())
        self.config.set('contractor_details', 'measuring_tools', self.measuring_tools_entry.get())

        with open('config.ini', 'w') as f:
            self.config.write(f)

        list_file = ['izo.xlsx', 'odgromy.xlsx', 'roznicowka.xlsx', 'zerowanie.xlsx']
        for file in list_file:
            wb = load_workbook(filename=file)
            ws = wb.active
            ws['A2'] = self.name_entry.get()
            ws['A2'].font = Font(bold=True)
            wb.save(file)

        wb = load_workbook(filename='strona_tytulowa.xlsx')
        ws = wb.active
        ws['B12'] = self.config.get('contractor_details', 'name')
        ws['B13'] = self.config.get('contractor_details', 'street_and_number')
        ws['B14'] = self.config.get('contractor_details', 'zip_code_and_city')
        ws['B15'] = self.config.get('contractor_details', 'measuring_tools')
        wb.save('strona_tytulowa.xlsx')

        self.ws_number[f'B{self.ws_number.max_row + 1}'] = self.register_number_entry.get()

        self.wb_number.save('rejestr.xlsx')
        self.destroy()

    def set_path(self):
        self.filepath = ctk.filedialog.askdirectory(
            initialdir=self.config.get('path', 'path') if self.config.get('path', 'path') != '' else os.path.abspath(
                '..'))
        self.config.set('path', 'path', self.filepath)
        self.filepath_label = ctk.CTkLabel(self, text=self.filepath)
        self.filepath_label.grid(row=4, column=1, columnspan=3, **self.params)
