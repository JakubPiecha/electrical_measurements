import datetime

import customtkinter as ctk
import os
from collections import defaultdict
from openpyxl import load_workbook
from tkcalendar import DateEntry

ctk.set_appearance_mode('dark')

ctk.set_default_color_theme('green')


class MainFrame(ctk.CTkScrollableFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)

        self.generator_object = ElectricalMeasurementsGenerate()

        self.params = {'padx': 8, 'pady': 8, 'sticky': 'w'}

        self.settings_button = ctk.CTkButton(self, text="Ustaw dane firmy", command=self.open_settings_window)
        self.settings_button.grid(row=0, column=0, **self.params)

        self.settings_window = None

        self.front_page = ctk.CTkCheckBox(self, text='Strona tytułowa', onvalue='front_page', offvalue=0)
        self.front_page.grid(row=1, column=0, **self.params)

        self.zeroing = ctk.CTkCheckBox(self,
                                       text='Zerowanie', onvalue='zeroing', offvalue=0, command=self.show_zeroing)
        self.zeroing.grid(row=1, column=1, **self.params)

        self.differential = ctk.CTkCheckBox(self, text='Różnicówka', onvalue='differential', offvalue=0,
                                            command=self.show_differential)
        self.differential.grid(row=1, column=2, **self.params)

        self.insulation = ctk.CTkCheckBox(self, text='Izolacje', onvalue='insulation', offvalue=0)
        self.insulation.grid(row=1, column=3, **self.params)

        self.reflexes = ctk.CTkCheckBox(self, text='Odgromy', onvalue='reflexes', offvalue=0,
                                        command=self.show_reflexes)
        self.reflexes.grid(row=1, column=4, **self.params)

        self.investment_data_label = ctk.CTkLabel(self, text='Dane obiektu wykonanych pomiarów:')
        self.investment_data_label.grid(row=2, column=0, **self.params)

        self.investment_zip_code_entry = ctk.CTkEntry(self, placeholder_text='Kod pocztowy')
        self.investment_zip_code_entry.grid(row=2, column=1, **self.params)

        self.investment_city_entry = ctk.CTkEntry(self, placeholder_text='Miejscowość')
        self.investment_city_entry.grid(row=2, column=2, **self.params)

        self.investment_street_and_number_entry = ctk.CTkEntry(self, width=180,
                                                               placeholder_text='ulica, nr domu/mieszkania')
        self.investment_street_and_number_entry.grid(row=2, column=3, **self.params)

        self.investment_name_entry = ctk.CTkEntry(self, placeholder_text='Nazwa Obiektu')
        self.investment_name_entry.grid(row=2, column=4, **self.params)

        self.investors_label = ctk.CTkLabel(self, text='Dane Zleceniodawcy pomiarów:')
        self.investors_label.grid(row=3, column=0, **self.params)

        self.investors_zip_code_entry = ctk.CTkEntry(self, placeholder_text='Kod pocztowy')
        self.investors_zip_code_entry.grid(row=3, column=1, **self.params)

        self.investors_city_entry = ctk.CTkEntry(self, placeholder_text='Miejscowość')
        self.investors_city_entry.grid(row=3, column=2, **self.params)

        self.investors_street_and_number_entry = ctk.CTkEntry(self, width=180,
                                                              placeholder_text='ulica, nr domu/mieszkania')
        self.investors_street_and_number_entry.grid(row=3, column=3, **self.params)

        self.investors_name_entry = ctk.CTkEntry(self, placeholder_text='Nazwa')
        self.investors_name_entry.grid(row=3, column=4, **self.params)

        self.date_label = ctk.CTkLabel(self, text='Data wykonania pomiarów:')
        self.date_label.grid(row=4, column=0, **self.params)

        self.date = DateEntry(self, selectmode='day', background='dark', foreground='grey',
                              selectbackground='grey', locale='pl')
        self.date.grid(row=4, column=1, **self.params)

        self.generate_button = ctk.CTkButton(self, text='Generuj pliki', command=self.generate)
        self.generate_button.grid(row=9, column=5, padx=10, pady=15, sticky='w')

        self.zeroing_frame = ZeroingFrame(master=self, height=200, width=1150, corner_radius=10, fg_color="gray20",
                                          label_text='Zerowanie', label_anchor='center')

        self.differential_frame = DifferentialFrame(master=self, width=1150, height=200, corner_radius=10,
                                                    fg_color="gray20", label_text='Różnicówka', label_anchor='center')

        self.reflexes_frame = ReflexesFrame(master=self, corner_radius=10, fg_color="gray20")

    def open_settings_window(self):
        if self.settings_window is None or not self.settings_window.winfo_exists():
            self.settings_window = SettingsWindow(self)
            self.settings_window.grab_set()
        else:
            self.settings_window.focus()

    def show_zeroing(self):
        if self.zeroing.get():
            self.zeroing_frame.grid(row=7, column=0, padx=5, pady=5, sticky="nsew", columnspan=6)
        else:
            self.zeroing_frame.grid_forget()

    def show_differential(self):
        if self.differential.get():
            self.differential_frame.grid(row=8, column=0, padx=5, pady=5, sticky="nsew", columnspan=6)
        else:
            self.differential_frame.grid_forget()

    def show_reflexes(self):
        if self.reflexes.get():
            self.reflexes_frame.grid(row=6, column=0, padx=5, pady=5, sticky="nsew", columnspan=6)
        else:
            self.reflexes_frame.grid_forget()

    def generate(self):
        obj = self.generator_object.measurement_data
        obj['investment_zip_code'] = self.investment_zip_code_entry.get()
        obj['investment_name'] = self.investment_name_entry.get()
        obj['investment_street_and_number'] = self.investment_street_and_number_entry.get()
        obj['investment_city'] = self.investment_city_entry.get()
        obj['investors_zip_code'] = self.investors_zip_code_entry.get()
        obj['investors_name'] = self.investors_name_entry.get()
        obj['investors_street_and_number'] = self.investors_street_and_number_entry.get()
        obj['investors_city'] = self.investors_city_entry.get()
        obj['date'] = self.date.get()
        obj.update(self.zeroing_frame.building_diagram)
        obj.update(self.differential_frame.building_diagram)
        obj.update(self.reflexes_frame.reflexes_data)

        print(obj)

        self.generator_object.create_directory()

        call_dict = {
            self.front_page.get(): self.generator_object.print_front_page,
            self.differential.get(): self.generator_object.differential,
            self.zeroing.get(): self.generator_object.zeroing,
            self.insulation.get(): self.generator_object.insulation,
            self.reflexes.get(): self.generator_object.reflexes
        }

        for key in call_dict.keys():
            if key:
                call_dict[key]()


class ModelFrame(ctk.CTkScrollableFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.params = {'padx': 5, 'pady': 5, 'sticky': 'w'}
        self.name = ''
        self.type_of_security = []
        self.protection_type = []
        self.In = []
        self.building_diagram = {self.name: defaultdict(defaultdict)}
        self.floors_data = self.building_diagram[self.name]

        self.storey_name_label = ctk.CTkLabel(self, text='Nazwa kondygnacji')
        self.storey_name_label.grid(row=1, column=0, **self.params)

        self.storey_name = ctk.CTkEntry(self, placeholder_text='np. Parter')
        self.storey_name.grid(row=1, column=1, **self.params)

        self.add_storey_button = ctk.CTkButton(self, text='Dodaj kondygnację',
                                               command=lambda: self.add_storey(self.storey_name.get()))
        self.add_storey_button.grid(row=2, column=1, **self.params)

    def add_storey(self, storey_name):
        last_row = self.grid_size()[-1]
        self.floors_data[storey_name] = defaultdict(dict)
        self.storey_name.delete(0, 'end')

        room_name_Label = ctk.CTkLabel(self, text=f'Nazwa/Liczba gniazdek dla - {storey_name}')
        room_name_Label.grid(row=last_row + 1, column=0, **self.params)

        room_name = ctk.CTkEntry(self, placeholder_text='Nazwa pomieszczenia')
        room_name.grid(row=last_row + 1, column=1, **self.params)

        power_sockets = ctk.CTkEntry(self, placeholder_text='Liczba gniazdek')
        power_sockets.grid(row=last_row + 1, column=2, **self.params)

        type_of_security_option = ctk.CTkOptionMenu(self, values=self.type_of_security)
        type_of_security_option.grid(row=last_row + 1, column=3, **self.params)

        protection_type_option = ctk.CTkOptionMenu(self, values=self.protection_type)
        protection_type_option.grid(row=last_row + 1, column=4, **self.params)

        In_option = ctk.CTkOptionMenu(self, values=self.In)
        In_option.grid(row=last_row + 1, column=5, **self.params)

        add_storey_button = ctk.CTkButton(self, text='Dodaj',
                                          command=lambda: self.add_room(storey_name, room_name, power_sockets,
                                                                        type_of_security_option, protection_type_option,
                                                                        In_option))
        add_storey_button.grid(row=last_row + 1, column=6, **self.params)

    def add_room(self, storey_name, room_name, power_sockets,
                 type_of_security_option, protection_type_option, In_option):
        self.floors_data[storey_name].update({
            room_name.get(): [power_sockets.get(), type_of_security_option.get(), protection_type_option.get(),
                              In_option.get()]
        })
        room_name.delete(0, 'end')
        power_sockets.delete(0, 'end')


class ZeroingFrame(ModelFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.name = 'Zeroing'
        self.building_diagram = {self.name: defaultdict(defaultdict)}
        self.floors_data = self.building_diagram[self.name]
        self.type_of_security = ['S301', 'S303', 'S191', 'S193', 'WT', 'BM']
        self.protection_type = ['B', 'A', 'C', 'D', 'F', 'gG']
        self.In = ['16', '2', '4', '6', '10', '20', '25', '32', '40', '50', '63', '80', '100', '125', '160', '315',
                   '400']


class DifferentialFrame(ModelFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.name = 'Differential'
        self.building_diagram = {self.name: defaultdict(defaultdict)}
        self.floors_data = self.building_diagram[self.name]
        self.type_of_security = ['P302', 'P304']
        self.protection_type = ['AC', 'A', 'AC-s', 'A-s']
        self.In = ['0,03', '0,006', '0,01', '0,1']


class ReflexesFrame(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.params = {'padx': 5, 'pady': 5, 'sticky': 'w'}

        self.reflexes_data = {}

        self.grounding_label = ctk.CTkLabel(self, text='Dane instalacji odgromowej')
        self.grounding_label.grid(row=1, column=0, **self.params)

        self.grounding = ctk.CTkEntry(self, placeholder_text='Liczba uziomów')
        self.grounding.grid(row=1, column=1, **self.params)

        self.ground_moisture = ctk.CTkOptionMenu(self, values=['1.4', '1.2', '2.0'])
        self.ground_moisture.grid(row=1, column=2, **self.params)

        self.required_resistance = ctk.CTkOptionMenu(self, values=['20', '1', '5', '7', '10', '30', '40', '50'])
        self.required_resistance.grid(row=1, column=3, **self.params)

        self.add_storey_button = ctk.CTkButton(self, text='Dodaj', command=self.add_reflexes_data)
        self.add_storey_button.grid(row=1, column=4, **self.params)

    def add_reflexes_data(self):
        self.reflexes_data.update(
            {'Reflexes': [self.grounding.get(), self.ground_moisture.get(), self.required_resistance.get()]})
        print(self.reflexes_data)


class SettingsWindow(ctk.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.params = {'padx': 8, 'pady': 8, 'sticky': 'nsew'}
        self.geometry("600x500")
        self.title('Dane Firmy Wykonującej Pomiary')

        self.name_label = ctk.CTkLabel(self, text='Nazwa firmy')
        self.name_label.grid(row=0, column=0, **self.params)

        self.name_entry = ctk.CTkEntry(self, placeholder_text='Nazwa Firmy')
        self.name_entry.grid(row=0, column=1, columnspan=3, **self.params)

        self.address_label = ctk.CTkLabel(self, text='Adres firmy')
        self.address_label.grid(row=1, column=0, **self.params)

        self.zip_code_and_city_entry = ctk.CTkEntry(self, width=200, placeholder_text='Kod pocztowy i miasto')
        self.zip_code_and_city_entry.grid(row=1, column=1, **self.params)

        self.street_and_number_entry = ctk.CTkEntry(self, width=200, placeholder_text='Ulica i numer')
        self.street_and_number_entry.grid(row=1, column=2, **self.params)

        self.measuring_tools_label = ctk.CTkLabel(self, text='Narzędzia pomiarów')
        self.measuring_tools_label.grid(row=2, column=0, **self.params)

        self.measuring_tools_entry = ctk.CTkEntry(self, placeholder_text='np. MIC-3 nr 345758')
        self.measuring_tools_entry.grid(row=2, column=1, **self.params, columnspan=3)

        self.register_number_label = ctk.CTkLabel(self, text='Numer rejestru')
        self.register_number_label.grid(row=3, column=0, **self.params)

        self.register_number_entry = ctk.CTkEntry(self, placeholder_text='np. 25')
        self.register_number_entry.grid(row=3, column=1, **self.params)

        self.set_and_save_button = ctk.CTkButton(self, text='Zapisz i zastosuj', command=self.save_data)
        self.set_and_save_button.grid(row=4, column=2, **self.params)

    def save_data(self):

        pass


class App(ctk.CTk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.title('Dane Pomiarów Elektrycznych')

        self.main_frame = MainFrame(master=self, width=1200, height=850, corner_radius=0, fg_color="transparent")
        self.main_frame.grid(row=0, column=0, sticky="nsew")


class ElectricalMeasurementsGenerate:
    def __init__(self):
        self.measurement_data = {
            'investment_zip_code': '',
            'investment_name': '',
            'investment_street_and_number': '',
            'investment_city': '',
            'investors_zip_code': '',
            'investors_name': '',
            'investors_street_and_number': '',
            'investors_city': '',
            'date': ''
        }
        self.path = ''

    def create_directory(self):
        # TODO ADD validation if directory exist and error massage
        investors_name = self.remove_special_char("_".join(self.measurement_data['investors_name'].split()))
        objects_name = self.remove_special_char("_".join(self.measurement_data['investment_name'].split()))
        objects_adres = self.remove_special_char(
            "_".join(self.measurement_data['investment_street_and_number'].split()))

        directory = f'{investors_name}_{objects_name}_{objects_adres}'
        self.path = os.path.join(os.path.abspath('..'), directory)
        os.mkdir(self.path)

    def save_file(self, filename, file):
        savefile = f'{filename.split(".")[0]}_{"_".join((self.measurement_data["investment_name"]).split())}.xlsx'
        file.save(os.path.join(self.path, savefile))

    def insert_column(self, wb, filename):
        row = 10
        ws = wb.active
        self.add_name(ws)
        if filename == 'zerowanie.xlsx':
            type_dict = self.measurement_data['Zeroing']
        else:
            type_dict = self.measurement_data['Differential']

        for name_floor, floor in type_dict.items():
            ws[f'C{str(row)}'] = name_floor
            row += 1

            for room, ps_data in floor.items():
                ws[f'C{str(row)}'] = room
                row += 1

                for ps in range(1, int(ps_data[0]) + 1):
                    ws[f'B{str(row)}'] = ps
                    ws[f'C{str(row)}'] = f'Gniazdo jednofazowe A/Z 230V nr {ps}'
                    ws[f'D{str(row)}'] = f'{ps_data[1]}'
                    ws[f'E{str(row)}'] = f'{ps_data[2]}'
                    ws[f'F{str(row)}'] = f'{ps_data[3]}'

                    if filename == 'zerowanie.xlsx':
                        ws[f'G{str(row)}'] = f'=IF(E{row}="A",F{row}*2.5,IF(E{row}="b",F{row}*5,IF(E{row}="c",' \
                                             f'F{row}*10,IF(E{row}="d",F{row}*20,IF(E{row}="F",' \
                                             f'F{row}*3,IF(E{row}="gG",F{row}*5,""))))))'
                        ws[f'I{str(row)}'] = f'=IF(AND(E{row}>0,F{row}>0,G{row}>0),184/G{row},"")'
                        ws[f'J{str(row)}'] = f'=IF(H{row}>0,IF(I{row}>H{row},"dobry","zły"),"")'
                    else:
                        ws[f'I{str(row)}'] = f'=IF(AND(E{row}>0,F{row}>0,G{row}>0),1667,"")'
                        ws[f'J{str(row)}'] = f'=IF(G{row}>0,IF(F{row}>G{row},"dobry","zły"),"")'
                    row += 1
        return wb

    def add_to_register(self):
        filename = 'rejestr.xlsx'
        wb = load_workbook(filename=filename)
        ws = wb.active
        max_row = wb.active.max_row
        year = self.measurement_data['date'].split('.')[-1]

        if ws[f'C{max_row}'].value == 'Data':
            number = ws[f'B{max_row + 1}'] = 1
        else:
            if isinstance(ws[f'C{max_row}'].value, datetime.datetime):
                last_year = str(ws[f'C{max_row}'].value.date()).split('-')[0]
            else:
                last_year = str(ws[f'C{max_row}'].value).split('.')[-1]

            number = ws[f'B{max_row + 1}'] = ws[f'B{max_row}'].value + 1 if last_year == year else 1

        ws[f'A{max_row + 1}'] = f'{self.measurement_data["investors_name"]} -' \
                                f' {self.measurement_data["investment_name"]} - ' \
                                f'{self.measurement_data["investment_street_and_number"]}'

        ws[f'C{max_row + 1}'] = self.measurement_data['date']

        wb.save('rejestr.xlsx')
        return f'{number}/{year}'

    def print_front_page(self):
        filename = 'strona_tytulowa.xlsx'
        wb = load_workbook(filename=filename)
        ws = wb.active
        ws['B6'] = self.measurement_data['investors_name']
        ws['B7'] = self.measurement_data['investors_street_and_number']
        ws['B8'] = self.measurement_data['investors_zip_code'] + ' ' + self.measurement_data['investment_city']
        ws['B9'] = self.measurement_data['investment_name']
        ws['B10'] = self.measurement_data['investment_street_and_number']
        ws['B11'] = self.measurement_data['investment_zip_code'] + ' ' + self.measurement_data['investment_city']
        ws['B17'] = self.measurement_data['date']
        date = self.measurement_data['date'].split('.')
        next_date = date[1] + "/" + (str(int(date[-1]) + 5))
        ws['B28'] = next_date
        ws['B2'] = self.add_to_register()

        self.save_file(filename, wb)

    def zeroing(self):
        filename = 'zerowanie.xlsx'
        wb = load_workbook(filename=filename)
        self.insert_column(wb, filename)

        self.save_file(filename, wb)

    def differential(self):
        filename = 'roznicowka.xlsx'
        wb = load_workbook(filename=filename)
        self.insert_column(wb, filename)
        self.save_file(filename, wb)

    def insulation(self):
        filename = 'izo.xlsx'
        wb = load_workbook(filename=filename)
        ws = wb.active
        self.add_name(ws)
        self.save_file(filename, wb)

    def reflexes(self):
        filename = 'odgromy.xlsx'
        wb = load_workbook(filename=filename)
        ws = wb.active
        self.add_name(ws)
        row = 10
        data = self.measurement_data['Reflexes']

        for ps in range(1, int(data[0]) + 1):
            ws[f'A{str(row)}'] = ps
            ws[f'B{str(row)}'] = ps
            ws[f'C{str(row)}'] = f'Uziom nr {ps}'
            ws[f'E{str(row)}'] = float(f'{data[1]}')
            ws[f'F{str(row)}'] = f'=IF(D{row}*E{row}>0,D{row}*E{row},"")'
            ws[f'G{str(row)}'] = int(f'{data[2]}')
            ws[f'H{str(row)}'] = f'=IF(D{row}>0,IF(F{row}<G{row},"dobry","zły"),"")'
            row += 1

        self.save_file(filename, wb)

    def add_name(self, ws):
        ws['A7'] = f'{self.measurement_data["investment_zip_code"]} {self.measurement_data["investment_city"]},' \
                   f' {self.measurement_data["investment_street_and_number"]} -' \
                   f' {self.measurement_data["investment_name"]}'
        return ws

    @staticmethod
    def remove_special_char(string):
        special_char = ['\\', '/', '[', ']', ':', '¦', '<', '>', '+', '=', ';', ',', '*', '?', '"']
        new_string = "".join(filter(lambda char: char not in special_char, string))
        return new_string


app = App()

app.mainloop()
