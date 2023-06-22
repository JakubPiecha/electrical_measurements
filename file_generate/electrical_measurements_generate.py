import configparser
import datetime
import os

from CTkMessagebox import CTkMessagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


class ElectricalMeasurementsGenerate:
    def __init__(self):
        self.measurement_data = {
            'investment_zip_code': '',
            'investment_name': '',
            'investment_street_and_number': '',
            'investment_city': '',
            'investors_zip_code': '',
            'investors_name': '',
            'investors_street_and_number': '',
            'investors_city': '',
            'date': ''
        }
        self.path = ''
        self.alignment = Alignment(horizontal='center', vertical='center')

    def create_directory(self):
        investors_name = self.remove_special_char("_".join(self.measurement_data['investors_name'].split()))
        objects_name = self.remove_special_char("_".join(self.measurement_data['investment_name'].split()))
        objects_adres = self.remove_special_char(
            "_".join(self.measurement_data['investment_street_and_number'].split()))
        year = self.measurement_data['date'].split('.')[-1]
        directory = f'{investors_name}_{objects_name}_{objects_adres}'
        config = configparser.ConfigParser()
        config.read('config.ini')
        self.path = config.get('path', 'path')

        if os.path.isdir(os.path.join(self.path, year)):
            self.path = os.path.join(self.path, year, directory)

            if os.path.isdir(self.path):
                confirm_box = CTkMessagebox(title='UWAGA',
                                            message='Katalog już istnieje!! '
                                                    'Kontynuacja może spowodować nadpisanie danych!!! '
                                                    'Czy chcesz kontynuować?',
                                            icon='question', option_1='Tak', option_2='Nie')
                if confirm_box.get() == 'Nie':
                    return False
            else:
                os.mkdir(self.path)
        else:
            os.mkdir(os.path.join(self.path, year))
            self.path = os.path.join(self.path, year, directory)
            os.mkdir(self.path)

        return True

    def save_file(self, filename, file):
        savefile = f'{filename.split(".")[0]}_{"_".join((self.measurement_data["investment_name"]).split())}.xlsx'
        file.save(os.path.join(self.path, savefile))

    def insert_column(self, wb, filename):
        row = 10
        ws = wb.active
        self.add_name(ws)

        if filename == 'zerowanie.xlsx':
            type_dict = self.measurement_data['Zeroing']
        else:
            type_dict = self.measurement_data['Differential']

        for name_floor, floor in type_dict.items():
            ws[f'C{str(row)}'] = name_floor
            ws[f'C{str(row)}'].alignment = self.alignment
            ws[f'C{str(row)}'].font = Font(bold=True)
            row += 1

            for room, ps_data in floor.items():
                if ps_data is not None:
                    ws[f'C{str(row)}'] = room
                    row += 1

                    for ps in range(1, int(ps_data[0]) + 1):
                        ws[f'B{str(row)}'] = ps
                        ws[f'B{str(row)}'].alignment = self.alignment
                        ws[f'C{str(row)}'] = f'Gniazdo jednofazowe A/Z 230V nr {ps}'
                        ws[f'C{str(row)}'].alignment = self.alignment
                        ws[f'D{str(row)}'] = f'{ps_data[1]}'
                        ws[f'D{str(row)}'].alignment = self.alignment
                        ws[f'E{str(row)}'] = f'{ps_data[2]}'
                        ws[f'E{str(row)}'].alignment = self.alignment
                        ws[f'F{str(row)}'] = float(ps_data[3])
                        ws[f'F{str(row)}'].alignment = self.alignment

                        if filename == 'zerowanie.xlsx':
                            ws[f'G{str(row)}'] = f'=IF(E{row}="A",F{row}*2.5,IF(E{row}="b",F{row}*5,IF(E{row}="c",' \
                                                 f'F{row}*10,IF(E{row}="d",F{row}*20,IF(E{row}="F",' \
                                                 f'F{row}*3,IF(E{row}="gG",F{row}*5,""))))))'
                            ws[f'G{str(row)}'].alignment = self.alignment
                            ws[f'I{str(row)}'] = f'=IF(AND(E{row}>0,F{row}>0,G{row}>0),184/G{row},"")'
                            ws[f'I{str(row)}'].alignment = self.alignment
                            ws[f'J{str(row)}'] = f'=IF(H{row}>0,IF(I{row}>H{row},"dobry","zły"),"")'
                            ws[f'J{str(row)}'].alignment = self.alignment
                        else:
                            ws[f'I{str(row)}'] = f'=IF(AND(E{row}>0,F{row}>0,G{row}>0),1667,"")'
                            ws[f'I{str(row)}'].alignment = self.alignment
                            ws[f'J{str(row)}'] = f'=IF(G{row}>0,IF(F{row}>G{row},"dobry","zły"),"")'
                            ws[f'J{str(row)}'].alignment = self.alignment
                        row += 1

        ws.print_area = f'A1:J{row}'

        return wb

    def add_to_register(self):
        filename = 'rejestr.xlsx'
        wb = load_workbook(filename=filename)
        ws = wb.active
        max_row = wb.active.max_row
        year = self.measurement_data['date'].split('.')[-1]

        if ws[f'C{max_row}'].value == 'Data':
            number = ws[f'B{max_row + 1}'] = 1
        else:
            if isinstance(ws[f'C{max_row}'].value, datetime.datetime):
                last_year = str(ws[f'C{max_row}'].value.date()).split('-')[0]
            else:
                last_year = str(ws[f'C{max_row}'].value).split('.')[-1]

            number = ws[f'B{max_row + 1}'] = ws[f'B{max_row}'].value + 1 if last_year == year else 1

        ws[f'A{max_row + 1}'] = f'{self.measurement_data["investors_name"]} -' \
                                f' {self.measurement_data["investment_name"]} - ' \
                                f'{self.measurement_data["investment_street_and_number"]}'

        ws[f'C{max_row + 1}'] = self.measurement_data['date']

        wb.save('rejestr.xlsx')
        return f'{number}/{year}'

    def print_front_page(self):
        filename = 'strona_tytulowa.xlsx'
        wb = load_workbook(filename=filename)
        ws = wb.active
        ws['B6'] = self.measurement_data['investors_name']
        ws['B7'] = self.measurement_data['investors_street_and_number']
        ws['B8'] = self.measurement_data['investors_zip_code'] + ' ' + self.measurement_data['investment_city']
        ws['B9'] = self.measurement_data['investment_name']
        ws['B10'] = self.measurement_data['investment_street_and_number']
        ws['B11'] = self.measurement_data['investment_zip_code'] + ' ' + self.measurement_data['investment_city']
        ws['B17'] = self.measurement_data['date']
        date = self.measurement_data['date'].split('.')
        next_date = date[1] + "/" + (str(int(date[-1]) + 5))
        ws['B28'] = next_date
        ws['B2'] = self.add_to_register()
        ws.print_area = 'A1:B39'

        self.save_file(filename, wb)

    def zeroing(self):
        filename = 'zerowanie.xlsx'
        wb = load_workbook(filename=filename)
        self.insert_column(wb, filename)

        self.save_file(filename, wb)

    def differential(self):
        filename = 'roznicowka.xlsx'
        wb = load_workbook(filename=filename)
        self.insert_column(wb, filename)
        self.save_file(filename, wb)

    def insulation(self):
        filename = 'izo.xlsx'
        wb = load_workbook(filename=filename)
        ws = wb.active
        self.add_name(ws)
        self.save_file(filename, wb)

    def reflexes(self):
        filename = 'odgromy.xlsx'
        wb = load_workbook(filename=filename)
        ws = wb.active
        self.add_name(ws)
        row = 10
        data = self.measurement_data['Reflexes']
        if data:
            for ps in range(1, int(data[0]) + 1):
                ws[f'A{str(row)}'] = ps
                ws[f'A{str(row)}'].alignment = self.alignment
                ws[f'B{str(row)}'] = ps
                ws[f'B{str(row)}'].alignment = self.alignment
                ws[f'C{str(row)}'] = f'Uziom nr {ps}'
                ws[f'C{str(row)}'].alignment = self.alignment
                ws[f'E{str(row)}'] = float(f'{data[1]}')
                ws[f'E{str(row)}'].alignment = self.alignment
                ws[f'F{str(row)}'] = f'=IF(D{row}*E{row}>0,D{row}*E{row},"")'
                ws[f'F{str(row)}'].alignment = self.alignment
                ws[f'G{str(row)}'] = int(f'{data[2]}')
                ws[f'G{str(row)}'].alignment = self.alignment
                ws[f'H{str(row)}'] = f'=IF(D{row}>0,IF(F{row}<G{row},"dobry","zły"),"")'
                ws[f'H{str(row)}'].alignment = self.alignment

                row += 1

        ws.print_area = f'A1:H{row}'
        self.save_file(filename, wb)



    def add_name(self, ws):
        ws['A7'] = f'{self.measurement_data["investment_zip_code"]} {self.measurement_data["investment_city"]},' \
                   f' {self.measurement_data["investment_street_and_number"]} -' \
                   f' {self.measurement_data["investment_name"]}'
        ws['A7'].font = Font(bold=True)
        return ws

    @staticmethod
    def remove_special_char(string):
        special_char = ['\\', '/', '[', ']', ':', '¦', '<', '>', '+', '=', ';', ',', '*', '?', '"']
        new_string = "".join(filter(lambda char: char not in special_char, string))
        return new_string
