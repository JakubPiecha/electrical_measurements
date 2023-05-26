from openpyxl import load_workbook, Workbook

dom = {
    'investment address': '43-190 Mikołów, ul. Jasna 142A',
    'investors': 'Jakub Pie',
    'number of floors': {
        'Piwnica': {
            'pom1': 1,
            'pom2': 5,
            'pom3': 3,
            'pom4': 3,
            'pom5': 2,
        },
        'Parter': {
            'pom1': 5,
            'pom2': 5,
            'pom3': 3,
            'pom4': 1,
            'pom5': 3,
        },
        'Poddasze': {
            'pom1': 5,
            'pom2': 8,
            'pom3': 3,
            'pom4': 1,
            'pom5': 3,
        }}}


def zeroing(dicts):
    wb = load_workbook(filename='zerowanie.xlsx')
    ws = wb.active
    ws['A7'] = dicts['investment address']
    row = 10
    column_c = 'C'
    for name, floors in dicts['number of floors'].items():
        ws[f'{column_c + str(row)}'] = name
        row += 1
        for room, number in floors.items():
            ws[f'{column_c + str(row)}'] = room
            row += 1
            for ps in range(1, number + 1):
                ws[f'{column_c + str(row)}'] = f'Gniazdo jednofazowe A/Z 230V nr {ps}'
                row += 1
    wb.save(f'zerowanie_{"_".join((dicts["investors"]).split())}.xlsx')


zeroing(dom)
